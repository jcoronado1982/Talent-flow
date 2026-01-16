import random
import time
import re
from playwright.sync_api import sync_playwright, Page

class JobSearchBrowser:
    def __init__(self, headless=False, user_data_dir="user_data"):
        self.playwright = sync_playwright().start()
        # Use persistent context to save session (cookies, login)
        # verify_downloads=False helps with some automation detection
        self.context = self.playwright.chromium.launch_persistent_context(
            user_data_dir=user_data_dir,
            headless=headless,
            executable_path="/usr/bin/google-chrome", # Use real Chrome for better stealth
            viewport={"width": 1280, "height": 800},
            # Stealth Args: REMOVED --no-sandbox as it triggers Google security warnings
            args=[
                "--disable-blink-features=AutomationControlled", 
            ],
            ignore_default_args=["--enable-automation"],
            java_script_enabled=True
        )
        
        if len(self.context.pages) > 0:
            self.page = self.context.pages[0]
        else:
            self.page = self.context.new_page()

        # 3. MAGIC: Inject Session from Real Chrome (The "Clone Token" strategy)
        try:
            import browser_cookie3
            import os
            print("   [Browser] Intentando clonar sesión de Google Chrome...")
            
            potential_paths = [
                os.path.expanduser("~/.config/google-chrome/Default/Cookies"),
                os.path.expanduser("~/.config/google-chrome/Profile 1/Cookies"),
                os.path.expanduser("~/.config/google-chrome/Profile 2/Cookies")
            ]
            
            cookies_to_add = []
            
            for path in potential_paths:
                if not os.path.exists(path): continue
                try:
                    # Extract ALL google.com cookies
                    cj_temp = browser_cookie3.chrome(cookie_file=path, domain_name=".google.com")
                    for c in cj_temp:
                         # Convert http.cookiejar.Cookie to Playwright dict
                         cookie_dict = {
                             "name": c.name,
                             "value": c.value,
                             "domain": c.domain,
                             "path": c.path,
                             "secure": c.secure,
                             # Expiration handling
                             "expires": c.expires if c.expires else -1
                         }
                         # Clean up optional fields that might break Playwright
                         # Playwright hates keys with None values sometimes or extra keys
                         
                         cookies_to_add.append(cookie_dict)
                    
                    if cookies_to_add:
                        print(f"   ✨ [Magic] Clonando {len(cookies_to_add)} cookies desde {path}")
                        self.context.add_cookies(cookies_to_add)
                        break
                except Exception as e:
                    # print(f" debug error cloning: {e}") 
                    continue

        except Exception as e:
            print(f"   [Browser] Warning: Could not clone session: {e}")

    def human_delay(self, min_seconds=1, max_seconds=2):
        """Simulates a human-like delay (Optimized for testing)."""
        time.sleep(random.uniform(min_seconds, max_seconds))

    def simulate_human_reading(self):
        """Simulates scrolling and mouse movements like a human reading."""
        try:
            # Random mouse moves (reduced)
            for _ in range(random.randint(1, 2)):
                x = random.randint(100, 1000)
                y = random.randint(100, 700)
                self.page.mouse.move(x, y)
                time.sleep(random.uniform(0.1, 0.2))
            
            # Scroll down slowly (faster)
            total_height = self.page.evaluate("document.body.scrollHeight")
            current_position = 0
            while current_position < total_height:
                scroll_step = random.randint(600, 1000) # Bigger steps
                current_position += scroll_step
                self.page.mouse.wheel(0, scroll_step)
                time.sleep(random.uniform(0.3, 0.7)) # Faster scroll
                # Stop if we scrolled too much (e.g., footer)
                if current_position > 2000: break
                
        except Exception as e:
            print(f"Warning in human simulation: {e}")

    def close(self):
        """Closes the browser."""
        self.context.close()
        self.playwright.stop()

    def get_page_content(self):
        """Returns the full HTML of the current page."""
        return self.page.content()

    def login(self, site, email, password):
        """Logs into the specified site."""
        print(f"Logging into {site}...")
        try:
            if site == "linkedin":
                self.page.goto("https://www.linkedin.com/login")
                self.human_delay()
                self.page.fill("#username", email)
                self.human_delay(0.5, 1.5)
                self.page.fill("#password", password)
                self.human_delay(0.5, 1.5)
                self.page.click("button[type='submit']")
                self.page.wait_for_load_state("networkidle")
                print("Login submitted.")
                
                # Manual 2FA Handling
                # Smart Wait: Detect Feed instead of asking user to press Enter
                print("   [Login] Verificando acceso al Feed (Esperando 2FA si es necesario)...")
                try:
                    # Wait up to 60s for user to solve 2FA or for page to load
                    self.page.wait_for_url("**/feed/**", timeout=60000)
                    print("   ✅ [Login] Feed detectado. Continuando...")
                except:
                    print("   ⚠️ [Login] No se detectó el Feed en 60s. Continuando bajo riesgo...")
                
                # Check if we are logged in (optional, but good for stability)
                # self.human_delay(3, 5) 
            elif site == "computrabajo":
                # Placeholder for Computrabajo
                pass
        except Exception as e:
            print(f"Error logging in: {e}")

    def search_jobs(self, site, query, location, time_filter="r259200"):
        """Performs a job search. time_filter: r86400(24h), r259200(3d), r604800(1w)."""
        print(f"Searching {site} for '{query}' in '{location}' (Filter: {time_filter})...")
        if site == "linkedin":
            # URL encoding might be needed for complex queries
            url = f"https://www.linkedin.com/jobs/search/?keywords={query}&location={location}&f_TPR={time_filter}"
            self.page.goto(url)
            self.human_delay(2, 3)
            # Wait for results to load
            try:
                self.page.wait_for_selector(".jobs-search-results-list", timeout=10000)
            except:
                print("Warning: Job list selector not found (might need manual interaction or layout changed)")

    def extract_job_links(self, site, limit=3):
        """Extracts job links from the search results."""
        links = []
        if site == "linkedin":
            # Select job cards. Note: Selectors vary greatly on LinkedIn (Logged in vs Public)
            # This targets the logged-in search results view
            try:
                # Use a broad strategy to find job links
                # .job-card-container__link is common in the list view
                link_elements = self.page.query_selector_all("a.job-card-container__link")
                
                for link_el in link_elements:
                    if limit and len(links) >= limit:
                        break
                    
                    href = link_el.get_attribute('href')
                    if href and "/jobs/view/" in href:
                         url = f"https://www.linkedin.com{href}" if href.startswith("/") else href
                         # Clean URL (remove tracking)
                         url = url.split("?")[0]
                         if url not in links:
                             links.append(url)
                
                if not links:
                    print("No links found with primary selector. Trying backup...")
                    # Backup for public view or different layout
                    link_elements = self.page.query_selector_all("a.base-card__full-link")
                    for link_el in link_elements:
                         if limit and len(links) >= limit: break
                         href = link_el.get_attribute('href')
                         if href:
                             links.append(href.split("?")[0])

            except Exception as e:
                print(f"Error extracting links: {e}")
        return links
    
    # Oops, extract_job_links is defined in this file, I shouldn't mess it up. 
    # Let's target get_job_details directly.

    
    def _extract_details_from_page(self):
        """Helper to extract details from the CURRENTLY visible page/pane."""
        details = {
            "description": "", 
            "date": "Unknown",
            "company": "Unknown",
            "location": "Unknown",
            "work_mode": "Unknown"
        }
        try:
             # 1. Extract Title
             try:
                 title_el = self.page.query_selector(".job-details-jobs-unified-top-card__job-title h1")
                 if not title_el: title_el = self.page.query_selector("h2.job-details-jobs-unified-top-card__job-title")
                 if not title_el: title_el = self.page.query_selector(".job-details-jobs-unified-top-card__job-title")
                 if title_el: details["title"] = title_el.inner_text().strip()
             except: pass

             # 1. Extract Date (Time Ago) & Location & Work Mode
             # STRATEGY: Grab the text of the entire Unified Top Card and use Regex.
             # This avoids issues with specific sub-selectors or separators.
             try:
                 # Try specific container first (Updated to match user screenshot: -container suffix)
                 top_card_el = self.page.query_selector(".job-details-jobs-unified-top-card__primary-description-container")
                 if not top_card_el:
                     top_card_el = self.page.query_selector(".job-details-jobs-unified-top-card__primary-description") 
                 if not top_card_el:
                     # Fallback to the whole top card container
                     top_card_el = self.page.query_selector(".job-details-jobs-unified-top-card")
                 
                 if top_card_el:
                     full_text = top_card_el.inner_text().replace("\n", " ").strip()
                     print(f"   [Debug] Top Card Text: {full_text[:100]}...") # Log first 100 chars
                     
                     # --- DATE REGEX ---
                     date_patterns = [
                         r"(\d+\s+(?:hour|minute|day|week|month)s?\s+ago)",
                         r"(just\s+now)",
                         r"(hace\s+\d+\s+(?:hora|minuto|día|semana|mes)s?)", 
                         r"(recién\s+publicado)",
                         r"(\d+\s+(?:h|d|w|m|y)\s+ago)"
                     ]
                     
                     # Priority 1: Check individual spans (tvm__text--low-emphasis) as shown in screenshot
                     spans = top_card_el.query_selector_all("span.tvm__text--low-emphasis")
                     for span in spans:
                         span_text = span.inner_text().strip()
                         for pat in date_patterns:
                             if re.search(pat, span_text, re.IGNORECASE):
                                 details["date"] = span_text # Found clean date in span
                                 print(f"   [Debug] Date found in span: {details['date']}")
                                 break
                         if details["date"] != "Unknown": break

                     # Priority 2: Use regex on full text if span lookup failed
                     if details["date"] == "Unknown":
                         for pat in date_patterns:
                             match = re.search(pat, full_text, re.IGNORECASE)
                             if match:
                                 details["date"] = match.group(1)
                                 break
                    
                     # --- WORK MODE REGEX ---
                     if re.search(r"\b(remote|remoto)\b", full_text, re.IGNORECASE): details["work_mode"] = "Remote"
                     elif re.search(r"\b(hybrid|híbrido)\b", full_text, re.IGNORECASE): details["work_mode"] = "Hybrid"
                     elif re.search(r"\b(on-site|presencial)\b", full_text, re.IGNORECASE): details["work_mode"] = "On-site"
                     
                     # --- WORK MODE FALLBACK (Pills/Insights) ---
                     if details["work_mode"] == "Unknown":
                         # Look for list items or spans that often contain these tags
                         insight_selectors = [
                             ".job-details-fit-level-preferences button", # Target from user screenshot
                             ".job-details-jobs-unified-top-card__job-insight",
                             ".job-details-jobs-unified-top-card__workplace-type",
                             "li.job-details-jobs-unified-top-card__job-insight",
                             ".ui-label",
                             ".mt2 span"
                         ]
                         for sel in insight_selectors:
                             els = self.page.query_selector_all(sel)
                             for el in els:
                                 txt = el.inner_text().lower().strip()
                                 # Clean up text (sometimes includes "Matches your profile" etc)
                                 if "remote" in txt or "remoto" in txt:
                                     details["work_mode"] = "Remote"
                                     break
                                 if "hybrid" in txt or "híbrido" in txt:
                                     details["work_mode"] = "Hybrid"
                                     break
                                 if "on-site" in txt or "presencial" in txt:
                                     details["work_mode"] = "On-site"
                                     break
                             if details["work_mode"] != "Unknown": break
                     
                     # --- LOCATION HEURISTIC ---
                     clean_text_for_split = re.sub(r"[·•|]", "###", full_text)
                     parts = [p.strip() for p in clean_text_for_split.split("###") if p.strip()]
                     for part in parts:
                         if details["date"] != "Unknown" and part in details["date"]: continue
                         if re.search(r"(applicant|solicitud|remote|remoto|hybrid|híbrido|onsite|presencial|ago|hace)", part, re.IGNORECASE): continue
                         details["location"] = part
                         break
                         
             except Exception as e:
                 print(f"Error parsing top card via Regex: {e}")

             # Fallback Date Selectors if regex failed
             if details["date"] == "Unknown":
                 date_fallback_selectors = [".tvm__text--low-emphasis", ".posted-time-ago__text"]
                 for ds in date_fallback_selectors:
                     el = self.page.query_selector(ds)
                     if el:
                         details["date"] = el.inner_text().strip()
                         break

             # 2. Extract Company
             try:
                 company_el = self.page.query_selector(".job-details-jobs-unified-top-card__company-name")
                 if not company_el: company_el = self.page.query_selector(".job-card-container__company-name")
                 if company_el: details["company"] = company_el.inner_text().strip()
             except: pass

             # 4. Extract Description
             selectors = [
                 ".jobs-description__content",  
                 "#job-details",               
                 ".show-more-less-html__markup", 
                 "article",                    
                 ".description"                
             ]
             
             found_el = None
             for selector in selectors:
                 try:
                      if self.page.is_visible(selector):
                          found_el = self.page.query_selector(selector)
                          if found_el: # break if found
                              break
                 except: continue

             if found_el:
                 try:
                     more_btn = found_el.query_selector("button[aria-label*='Show more']")
                     if more_btn and more_btn.is_visible():
                          more_btn.click()
                          self.human_delay(0.5,1)
                 except: pass
                 
                 details["description"] = found_el.inner_text()
                 
                 # --- EXTRACT RAW REQUIREMENTS SECTION ---
                 # Try to find the specific block for "Requirements" or "Experience"
                 # to replace the generic AI analysis
                 req_regex = r"(?i)(?:Requisitos|Requirements|Perfil|Profile|What you need|Who you are|Experiencia|Experience|Qualifications)(?:[\s:]+)(.*?)(?:Benefits|Beneficios|Ofrecemos|Offer|About|Sobre|Compensation|What we offer|TalentFlow|$)"
                 match_req = re.search(req_regex, details["description"], re.DOTALL)
                 if match_req:
                     raw_req = match_req.group(1).strip()
                     # If too short, might be a false positive title, keep full description
                     if len(raw_req) > 50:
                         details["raw_requirements"] = raw_req[:1000] # Cap at 1000 chars
                     else:
                         details["raw_requirements"] = details["description"][:1000]
                 else:
                     details["raw_requirements"] = details["description"][:1000]
             
             # Debug Check
             if not details["description"] or len(details["description"]) < 100:
                 print("   [Browser] Warning: Description empty or too short.")

        except Exception as e:
            print(f"Error extracting details helper: {e}")
            
        return details

    def get_job_details(self, site, url):
        """Extracts job description text and date."""
        print(f"Getting details for: {url}")
        details = {"description": "", "date": "Unknown"}
        
        if site == "linkedin":
            try:
                self.page.goto(url)
                self.human_delay()
                self.simulate_human_reading()
                details = self._extract_details_from_page()

            except Exception as e:
                print(f"Error getting details: {e}")
                
        return details

    def scan_search_results(self, site, limit, callback_fn):
        """
        Iterates through the search results list, clicking each job, 
        and extracting details from the right pane without leaving the page.
        """
        print(f"Scanning search results (Limit: {limit})...")
        if site != "linkedin":
            print("Scan only supported for LinkedIn currently.")
            return

        count_processed = 0
        try:
             # Wait for the list to appear
             list_selector = ".jobs-search-results-list"
             found_list = False
             try:
                 self.page.wait_for_selector(list_selector, timeout=5000)
                 found_list = True
             except:
                 # Fallbacks
                 if self.page.is_visible("ul.scaffold-layout__list-container"):
                     list_selector = "ul.scaffold-layout__list-container"
                     found_list = True
                 elif self.page.is_visible(".jobs-search__results-list"):
                     list_selector = ".jobs-search__results-list"
                     found_list = True
             
             print(f"   [Scan] List container strategy: {list_selector if found_list else 'Global Search'}")
             
             # Define card selector based on whether we found the list or not
             if found_list:
                 job_card_selector = f"{list_selector} li"
             else:
                 # Fallback: standard card container
                 job_card_selector = ".job-card-container" 
             
             # Loop
             # We use a while loop with re-querying
             index = 0
             while count_processed < limit:
                 # Re-query list items every time because DOM might update
                 cards = self.page.query_selector_all(job_card_selector)
                 
                 if index >= len(cards):
                     print("   [Scan] Reached end of visible list.")
                     # TODO: Scroll down to load more?
                     # For now, just scroll the list container if possible
                     try:
                         self.page.evaluate("document.querySelector('.jobs-search-results-list').scrollBy(0, 500)")
                         self.human_delay(1, 2)
                         cards = self.page.query_selector_all(job_card_selector)
                         if index >= len(cards): break # No new items loaded
                     except:
                         break

                 card = cards[index]
                 
                 # Scroll card into view
                 try:
                     card.scroll_into_view_if_needed()
                 except: pass

                 # Click it
                 try:
                     # Find the clickable target inside the card (usually the title or the card itself)
                     # Clicking the card itself usually works
                     # We get the Job ID or URL from the card anchor for reference
                     link_el = card.query_selector("a.job-card-container__link")
                     job_url = "Unknown"
                     if link_el:
                         href = link_el.get_attribute("href")
                         if href: 
                             job_url = href.split("?")[0]
                             if job_url.startswith("/"):
                                 job_url = f"https://www.linkedin.com{job_url}"
                     
                     print(f"   [Scan] Clicking job {index+1}/{limit}: {job_url}")
                     
                     # Click wrapper or link? Try clicking the container div inside the LI
                     clickable = card.query_selector("div.job-card-container")
                     if clickable: clickable.click()
                     else: card.click()
                     
                     self.human_delay(1.5, 2) # Wait for right pane to load
                     
                     # Extract details from right pane
                     details = self._extract_details_from_page()
                     
                     # Callback
                     # We assume the URL matches the one we clicked
                     if job_url == "Unknown": hash(details['description']) # fallback ID?

                     should_continue = callback_fn(details, job_url)
                     if should_continue is False:
                         print("   [Scan] Callback requested stop.")
                         break
                     
                     count_processed += 1
                     index += 1

                 except Exception as e:
                     print(f"   [Scan] Error processing card {index}: {e}")
                     index += 1
                     continue

        except Exception as e:
            print(f"Error during scan: {e}")
            
        return count_processed

    def click_like_an_ai(self):
        """
        Simula la lógica de un Agente AI (como Perplexity):
        1. Busca por Semántica (Accesibilidad) -> Lo que usan los ciegos (LinkedIn no puede cambiar esto).
        2. Busca por Texto Visual.
        3. Busca por Selectores CSS clásicos.
        4. Inyección de JavaScript (Fuerza bruta).
        Returns True if clicked, False otherwise.
        """
        print("\n   🤖 [Browser] Iniciando protocolo de clic inteligente...")

        # PALABRAS CLAVE (Multilenguaje)
        # Regex para capturar: "Solicitar", "Solicitud sencilla", "Apply", "Easy Apply"
        pattern = re.compile(r"(solicitar|apply|sencilla|now)", re.IGNORECASE)

        # ---------------------------------------------------------
        # ESTRATEGIA 1: Semántica (Accessibility Tree) - LA MEJOR
        # ---------------------------------------------------------
        print("      1️⃣  Intentando búsqueda Semántica (Accessibility Role)...")
        try:
            # Busca un elemento que SEA un botón y que SE LLAME como el patrón
            btn = self.page.get_by_role("button", name=pattern).first
            
            if btn.is_visible():
                print(f"         ✨ ¡Encontrado! Texto: '{btn.text_content().strip()}'")
                print("         🖱️  Haciendo clic semántico...")
                btn.click(timeout=3000)
                return True
        except Exception as e:
            print(f"         ⚠️  Semántica falló: {e}")

        # ---------------------------------------------------------
        # ESTRATEGIA 2: Texto Visual (Lo que ve el humano)
        # ---------------------------------------------------------
        print("      2️⃣  Intentando búsqueda por Texto Visual...")
        try:
            text_btn = self.page.get_by_text(pattern).first
            if text_btn.is_visible():
                print("         🖱️  Haciendo clic en texto...")
                text_btn.click(force=True)
                return True
        except:
            pass

        # ---------------------------------------------------------
        # ESTRATEGIA 3: Selectores CSS (Legacy / Backup)
        # ---------------------------------------------------------
        print("      3️⃣  Intentando Selectores CSS clásicos...")
        selectors = [
            ".jobs-apply-button",
            ".jobs-s-apply button",
            "button[aria-label*='Apply']",
            ".jobs-apply-button--top-card button"
        ]
        for sel in selectors:
            if self.page.is_visible(sel):
                print(f"         🎯 Selector encontrado: {sel}")
                self.page.locator(sel).first.click()
                return True

        # ---------------------------------------------------------
        # ESTRATEGIA 4: Inyección JS (Opción Nuclear)
        # ---------------------------------------------------------
        print("      ☢️  Intentando Inyección Directa de JS (Bypass UI)...")
        result = self.page.evaluate("""
            () => {
                const xpath = "//button[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'apply') or contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'solicitar')]";
                const btn = document.evaluate(xpath, document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;
                if (btn) {
                    btn.click();
                    return true;
                }
                return false;
            }
        """)
        
        if result:
            print("         ✅ JS Click ejecutado con éxito.")
            return True

        print("      ❌ [Browser] No se pudo hacer clic con ninguna estrategia.")
        return False

    def create_google_sheet(self, report_data, google_email=None, google_password=None, output_filename=None):
        """Creates a local Excel report (and optionally uploads to Sheets)."""
        if not report_data: return
        
        # --- LOCAL EXCEL DECISION ---
        try:
             import openpyxl
             import os
             from openpyxl.styles import Font, PatternFill, Alignment
             from datetime import datetime
             
             # Format: report_1_13_01_2026_10_31.xlsx (Template)
             template_path = os.path.abspath("report_1_13_01_2026_10_31.xlsx")
             
             # Output filename
             if not output_filename:
                 timestamp = datetime.now().strftime('%d_%m_%Y_%H_%M')
                 output_filename = f"reports/report_FILLED_{timestamp}.xlsx"
             
             output_path = os.path.abspath(output_filename)
             
             # Ensure reports dir exists
             os.makedirs(os.path.dirname(output_path), exist_ok=True)
             
             if os.path.exists(template_path):
                 print(f"   [Excel] Loading template: {template_path}")
                 wb = openpyxl.load_workbook(template_path)
                 ws = wb.active
             else:
                 print("   [Excel] Template not found. Creating new workbook.")
                 wb = openpyxl.Workbook()
                 ws = wb.active
                 ws.title = "Job Analysis"
                 # Headers (Updated)
                 # Headers (Updated)
                 headers = ["Priority", "Match %", "Company", "Role", "Location", "Work Mode", "Date", "Source", "URL", "Requirements"]
                 ws.append(headers)
                 
                 # Style Headers
                 header_fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
                 header_font = Font(bold=True, color="FFFFFF", size=11)
                 for col_num, header in enumerate(headers, 1):
                     cell = ws.cell(row=1, column=col_num)
                     cell.fill = header_fill
                     cell.font = header_font
                     cell.alignment = Alignment(horizontal="center")
             
             # Rows
             for item in report_data:
                analysis = item['analysis']
                # MAPPING TO NEW HEADERS: 
                # [Priority, Match %, Company, Role, Location, Work Mode, Date, Source, URL, Requirements]
                ws.append([
                    int(analysis.get('priority_score', 4)), 
                    f"{analysis.get('match_percentage', 0)}%", 
                    item.get('company', 'Unknown'),
                    item.get('role', 'Unknown'), # Extracted Role or Search Term fallback
                    item.get('location', 'Unknown'),
                    item.get('work_mode', 'Unknown'),
                    item.get('date', 'Unknown'),
                    item['source'],
                    item['url'],
                    item.get('raw_requirements', analysis.get('analysis', '')).replace("\n", " ")[:3000] # Limit cell size
                ])
             
             # Auto-adjust columns (Basic)
             for col in ws.columns:
                 max_length = 0
                 column = col[0].column_letter # Get the column name
                 for cell in col:
                     try:
                         if len(str(cell.value)) > max_length:
                             max_length = len(str(cell.value))
                     except: pass
                 adjusted_width = (max_length + 2)
                 if adjusted_width > 50: adjusted_width = 50 # Cap width
                 ws.column_dimensions[column].width = adjusted_width

             wb.save(output_path)
             print(f"\n✅ REPORTE EXCEL CREADO: {output_path}")
             print("   (Formato: report_1_DD_MM_YYYY_HH_MM.xlsx)")
             
        except ImportError:
             print("   [Error] openpyxl not installed. Skipping Excel creation.")
        except Exception as e:
             print(f"   [Error] Excel creation failed: {e}")

        # --- GOOGLE SHEETS UPLOAD (DISABLED TEMPORARILY) ---
        upload_to_sheets = False # User requested to disable but keep logic
        
        if upload_to_sheets:
            print(f"Creating Google Sheet for {len(report_data)} jobs (Upload Strategy)...")
            # ... (Existing upload logic logic preserved below if we needed it) ...
            pass
        
        return # Stop here for now
        
        # 1. Generate Local CSV
        import csv
        import os
        
        csv_path = os.path.abspath("job_report.csv")
        try: 
            # ... (Rest of old CSV logic) ...
            self.page.wait_for_load_state("networkidle")
            self.human_delay(2, 3)
            
            # --- GOOGLE LOGIN HANDLING (If needed again) ---
            if "signin" in self.page.url or "accounts.google" in self.page.url:
                 # ... (Reuse login logic if safe, or rely on cookie injection) ...
                 pass 

            # 3. Trigger File Upload
            print("   [Sheets] Opening File Picker...")
            try:
                # Click the "Open file picker" folder icon
                # Selector strategy: The folder icon usually has a specific aria-label
                # Note: This might change, but it's standard for now.
                self.page.click("div[aria-label='Open file picker']", timeout=5000)
                self.human_delay(1, 2)
                
                # Click "Upload" tab
                # Use text selector for robustness across locales if possible, or reliable struct
                # "Upload" is usually the 2nd or 3rd tab.
                # Trying to find the tab by content "Upload" or "Subir"
                upload_tab = self.page.query_selector("div[role='tab']:has-text('Upload')")
                if not upload_tab:
                     upload_tab = self.page.query_selector("div[role='tab']:has-text('Subir')")
                
                if upload_tab:
                    upload_tab.click()
                    self.human_delay(1)
                    
                    # 4. Upload File
                    print("   [Sheets] Uploading CSV...")
                    # Expecting a file input of type file
                    # Once we select the file, Google Sheets usually auto-opens it
                    
                    # Some upload dialogs have a 'Select a file from your device' button wrapping the input
                    # We can use set_input_files on the page, Playwright finds the distinct file input
                    try:
                        with self.page.expect_file_chooser() as fc_info:
                             # Sometimes we need to click "Browse" to trigger chooser, 
                             # but set_input_files usually works on the input directly if present.
                             # Let's try locating the input.
                             file_input = self.page.query_selector("input[type='file']")
                             if file_input:
                                 file_input.set_input_files(csv_path)
                             else:
                                 # Or click the "Browse" button first?
                                 browse_btn = self.page.query_selector("text='Browse'")
                                 if not browse_btn: browse_btn = self.page.query_selector("text='Explorar'")
                                 
                                 if browse_btn:
                                     # This triggers OS dialog, handled by set_input_files? 
                                     # Playwright requires expect_file_chooser BEFORE action.
                                     # Actually, file_input.set_input_files is the modern way.
                                     pass
                    except:
                        # Fallback: Just try to set input files on the whole page frame if input is hidden
                        self.page.set_input_files("input[type='file']", csv_path)
                    
                    print("   [Sheets] Upload started. Waiting for sheet...")
                    self.human_delay(5, 8) # Wait for processing/redirect
                    
                else:
                    print("   [Warning] Could not find Upload tab.")

            except Exception as e:
                 print(f"   [Sheets] Upload automation failed: {e}")
                 print(f"   [Manual] Please upload '{csv_path}' manually.")
                 # Open the CSV locally as fallback so user sees SOMETHING
                 # self.page.goto(f"file://{csv_path}") 

            print("\n" + "="*50)
            print("✅ REPORTE LISTO")
            print(f"Archivo local: {csv_path}")
            print("Si no se abrió automáticamente, por favor súbelo a Sheets.")
            print("="*50 + "\n")
            
            time.sleep(5)
            
        except Exception as e:
            print(f"Error in upload process: {e}")
