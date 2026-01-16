import time
import re
import os
import shutil
import openpyxl
import sys

# Add project root to path
sys.path.append(os.path.abspath(os.path.dirname(__file__)))
from browser import JobSearchBrowser

# --- CONFIGURACIÓN DE PRUEBA ---
REPORT_PATH = "reports/report_FINAL_12_01_2026_19_02.xlsx"

def get_urls_from_report(path):
    """Extrae las URLs de la columna 'URL' o inferida del Excel."""
    urls = []
    if not os.path.exists(path):
        print(f"❌ Error: No se encuentra el reporte en {path}")
        return urls
        
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        
        # Identificar columna URL
        url_col_idx = None
        headers = [cell.value for cell in ws[1]]
        
        for idx, header in enumerate(headers):
            if header and isinstance(header, str) and ("url" in header.lower() or "enlace" in header.lower()):
                url_col_idx = idx
                break
        
        if url_col_idx is None:
            # Fallback: Asumir columna 9 (I) (índice 8) si no se encuentra header
            # O buscar la primera que parezca una URL
            print("⚠️ Header 'URL' no encontrado. Buscando primera columna con 'http'...")
            for row in ws.iter_rows(min_row=2, max_row=5):
                for idx, cell in enumerate(row):
                    if cell.value and isinstance(cell.value, str) and "http" in cell.value:
                        url_col_idx = idx
                        print(f"   ℹ️  Detectado URL en columna {idx+1}")
                        break
                if url_col_idx is not None: break
        
        if url_col_idx is None:
            print("❌ No se pudo identificar la columna de URLs.")
            return []

        # Extraer URLs
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[url_col_idx]:
                urls.append(row[url_col_idx])
                
        print(f"📋 Se cargaron {len(urls)} URLs del reporte.")
        return urls
        
    except Exception as e:
        print(f"Error leyendo Excel: {e}")
        return []

def run_test():
    urls = get_urls_from_report(REPORT_PATH)
    if not urls:
         print("⚠️ No hay URLs para procesar.")
         return

    # --- CLONAR SESIÓN (Manual para el test, aunque browser.py intenta clonar cookies, 
    # queremos el perfil completo de chrome si es posible, pero JobSearchBrowser usa user_data_dir) ---
    SOURCE_USER_DATA = "user_data"
    TEST_USER_DATA = f"user_data_test_click_{int(time.time())}"
    
    print(f"🔄 Preparando entorno de prueba...")
    # No need to clean up old one if we use unique name, but maybe clean up previous ones?
    # For now, just generate new one.
    if os.path.exists(TEST_USER_DATA):
        try:
             shutil.rmtree(TEST_USER_DATA)
        except: pass # Ignore if we can't delete self (unlikely if unique)
        
    if os.path.exists(SOURCE_USER_DATA):
        print(f"   📋 Clonando sesión activa desde '{SOURCE_USER_DATA}'...")
        try:
            # Ignorar archivos de bloqueo para evitar crashes
            def ignore_locks(dir, files):
                return [f for f in files if f.startswith("Singleton") or f == "Lock"]
            
            shutil.copytree(SOURCE_USER_DATA, TEST_USER_DATA, ignore=ignore_locks)
            print("   ✅ Sesión clonada exitosamente.")
        except Exception as e:
             print(f"   ⚠️ Error clonando sesión: {e}")
             print("   ⚠️ Se iniciará con un perfil limpio.")
    else:
         print("   ⚠️ No se encontró sesión previa ('user_data'). Se iniciará limpio.")

    print("🚀 Iniciando JobSearchBrowser...")
    # Instantiate the wrapper class
    browser = JobSearchBrowser(headless=False, user_data_dir=TEST_USER_DATA)
    
    try:
        print("🌍 Navegando a LinkedIn...")
        
        # Validar Login una vez
        browser.page.goto("https://www.linkedin.com/feed/")
        if "login" in browser.page.url or "signup" in browser.page.url:
             print("\n⚠️  POR FAVOR, LOGUEATE MANUALMENTE.")
             input("👉 Presiona ENTER cuando estés en el Feed...")
        
        # PROCESAR URLs
        print(f"\n🚀 Iniciando procesamiento de {len(urls)} ofertas...")
        
        for i, url in enumerate(urls):
            print(f"\n[{i+1}/{len(urls)}] Procesando: {url}")
            try:
                browser.page.goto(url)
                time.sleep(3) # Esperar carga inicial
                
                # Ejecutar lógica usando el método de clase integrado
                success = browser.click_like_an_ai()
                
                if success:
                    print(f"   ✅ [Postulación Abierta] Oferta {i+1} lista.")
                    time.sleep(2) 
                else:
                    print(f"   ❌ [Falló] No se pudo abrir modal.")

            except Exception as e:
                print(f"   ⚠️ Error procesando URL: {e}")
                
            time.sleep(1) # Pausa entre ofertas

    except Exception as e:
        print(f"❌ Error fatal en el test: {e}")
    finally:
        print("\n🏁 Procesamiento finalizado.")
        input("Presiona ENTER para cerrar el navegador...")
        browser.close()

if __name__ == "__main__":
    run_test()
